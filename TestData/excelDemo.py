import openpyxl

#1st we load the excel file
book = openpyxl.load_workbook(r"C:\Users\Araz_Markosian\Desktop\Selenium-Python-Automation-Testing-from-Scratch-and-Frameworks-master\PythonSelfFramework\PythonDemo.xlsx")
#let's hold the active sheet with the data
sheet = book.active
#read a single cell in the sheet
cell = sheet["B1"]
print(cell.value)
#write into a single cell in the sheet
sheet["B2"].value = "Araz"
print(sheet["B2"].value)
#now save this to your book
#$book.save("PythonDemo.xlsx")
#get the number of rows in your sheet
print(sheet.max_row)
#get the number of cols in your sheet
print(sheet.max_column)

#another way of accessing values in cells
print(sheet.cell(row = 1, column = 2).value)
#let's print all the cells in this sheet
for i in range(1, sheet.max_row + 1):
    for j in range(1, sheet.max_column + 1):
        print(sheet.cell(row=i, column=j).value)

#let's change the loop to print only col 1
print("all the values in col 1")
for i in range(1, sheet.max_row +1):
    print(sheet.cell(row = i, column = 1).value)

